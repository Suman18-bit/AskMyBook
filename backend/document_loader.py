"""
Multi-format document ingestion.

Your original CLI only loaded PDFs and text. This module extends that
to .docx and .xlsx, and routes each extension to the right LangChain
loader (or a lightweight custom loader for xlsx, since LangChain's
default Excel loader pulls in unstructured/openpyxl in ways that are
often heavier than needed here).
"""
from pathlib import Path

from langchain_community.document_loaders import PyPDFLoader, TextLoader, Docx2txtLoader
from langchain_core.documents import Document
import openpyxl


def _load_xlsx(path: str) -> list[Document]:
    """
    Converts each worksheet row into readable text so the RAG pipeline
    can retrieve over tabular data. Each row becomes 'col_a: val_a | col_b: val_b'
    joined with the sheet name, which embeds far better than raw cell dumps.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    documents = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        body_rows = rows[1:] if len(rows) > 1 else []

        for row_idx, row in enumerate(body_rows, start=2):
            pairs = [
                f"{headers[i]}: {cell}"
                for i, cell in enumerate(row)
                if cell is not None and i < len(headers)
            ]
            if not pairs:
                continue
            text = f"[Sheet: {sheet_name}, Row: {row_idx}] " + " | ".join(pairs)
            documents.append(
                Document(
                    page_content=text,
                    metadata={"source": path, "sheet": sheet_name, "row": row_idx},
                )
            )

    return documents


def load_document(file_path: str) -> list[Document]:
    """
    Loads a single document into LangChain Document objects, dispatching
    on file extension. Raises ValueError for unsupported types so callers
    get a clear error instead of a silent no-op.
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".pdf":
        loader = PyPDFLoader(file_path)
        return loader.load()

    elif ext == ".txt":
        loader = TextLoader(file_path, encoding="utf-8")
        return loader.load()

    elif ext == ".docx":
        loader = Docx2txtLoader(file_path)
        return loader.load()

    elif ext == ".xlsx":
        return _load_xlsx(file_path)

    else:
        raise ValueError(
            f"Unsupported file type: {ext}. Allowed: .pdf, .txt, .docx, .xlsx"
        )